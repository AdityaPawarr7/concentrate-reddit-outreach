"""Export graded posts to CSV and Excel."""

from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any

from grader import GradeResult
from outreach_templates import template_for_angle
from scraper import RedditPost

COLUMNS = [
    "scraped_at",
    "post_id",
    "subreddit",
    "title",
    "author",
    "created_utc",
    "score_reddit",
    "num_comments",
    "permalink",
    "alignment_score",
    "outreach_priority",
    "suggested_angle",
    "suggested_reply",
    "matched_signals",
    "rationale",
    "llm_graded",
    "selftext_preview",
]


def _row(post: RedditPost, grade: GradeResult) -> dict[str, Any]:
    return {
        "scraped_at": post.scraped_at.isoformat(),
        "post_id": post.post_id,
        "subreddit": post.subreddit,
        "title": post.title,
        "author": post.author,
        "created_utc": post.created_utc.isoformat(),
        "score_reddit": post.score,
        "num_comments": post.num_comments,
        "permalink": post.permalink,
        "alignment_score": grade.alignment_score,
        "outreach_priority": grade.outreach_priority,
        "suggested_angle": grade.suggested_angle,
        "suggested_reply": template_for_angle(grade.suggested_angle),
        "matched_signals": grade.matched_signals,
        "rationale": grade.rationale,
        "llm_graded": grade.llm_used,
        "selftext_preview": (post.selftext or "")[:500],
    }


def _load_existing_ids(path: Path) -> set[str]:
    if not path.exists():
        return set()
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return {row["post_id"] for row in reader if row.get("post_id")}


def export_csv(
    rows: list[dict[str, Any]],
    path: str | Path,
    *,
    append: bool = True,
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    existing_ids = _load_existing_ids(path) if append else set()
    new_rows = [r for r in rows if r["post_id"] not in existing_ids]
    if not new_rows and append:
        return path

    write_header = not path.exists() or not append
    mode = "w" if not append else "a"
    with path.open(mode, encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        if write_header:
            writer.writeheader()
        writer.writerows(new_rows)
    return path


def export_xlsx(all_rows: list[dict[str, Any]], path: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reddit Leads"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    sorted_rows = sorted(all_rows, key=lambda r: int(r.get("alignment_score") or 0), reverse=True)
    for row in sorted_rows:
        ws.append([row.get(c, "") for c in COLUMNS])

    priority_col = COLUMNS.index("outreach_priority") + 1
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=priority_col).value
        fill = None
        if val == "high":
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif val == "medium":
            fill = PatternFill("solid", fgColor="FFEB9C")
        if fill:
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    for col in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions[get_column_letter(COLUMNS.index("permalink") + 1)].width = 40

    wb.save(path)
    return path


def _read_all_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def export_all(
    graded: list[tuple[RedditPost, GradeResult]],
    *,
    csv_path: str | None = None,
    xlsx_path: str | None = None,
) -> dict[str, Path]:
    csv_path = Path(csv_path or os.getenv("OUTPUT_CSV", "./output/reddit_leads.csv"))
    xlsx_path = Path(xlsx_path or os.getenv("OUTPUT_XLSX", "./output/reddit_leads.xlsx"))

    rows = [_row(p, g) for p, g in graded]
    export_csv(rows, csv_path, append=True)

    all_rows = _read_all_csv(csv_path)
    batch_by_id = {r["post_id"]: r for r in rows}
    merged: dict[str, dict[str, Any]] = {r["post_id"]: r for r in all_rows}
    merged.update(batch_by_id)

    export_xlsx(list(merged.values()), xlsx_path)
    return {"csv": csv_path, "xlsx": xlsx_path}
